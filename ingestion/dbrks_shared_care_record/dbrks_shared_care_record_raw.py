# Databricks notebook source
#!/usr/bin python3

# -------------------------------------------------------------------------
# Copyright (c) 2021 NHS England and NHS Improvement. All rights reserved.
# Licensed under the MIT License. See license.txt in the project root for
# license information.
# -------------------------------------------------------------------------

"""
FILE:           dbrks_shared_care_record_raw.py
DESCRIPTION:
                Databricks notebook with code to append new raw data to historical
                data for the NHSX Analyticus unit metrics within the Shared Care Record topic.
                
USAGE:
                ...
CONTRIBUTORS:   Craig Shenton, Mattia Ficarelli, Chris Todd
CONTACT:        data@nhsx.nhs.uk
CREATED:        08th of Sept. 2022
VERSION:        0.0.2
"""

# COMMAND ----------

# Install libs
# -------------------------------------------------------------------------
%pip install geojson==2.5.* tabulate requests pandas pathlib azure-storage-file-datalake  beautifulsoup4 numpy urllib3 lxml dateparser regex openpyxl==3.1.0  pyarrow==5.0.*

# COMMAND ----------

# Imports
# -------------------------------------------------------------------------
# Python:
import os
import io
import tempfile
from datetime import datetime
import json
import regex as re

# 3rd party:
import pandas as pd
import numpy as np
import openpyxl
from pathlib import Path
from azure.storage.filedatalake import DataLakeServiceClient
from openpyxl import load_workbook
import collections

# Connect to Azure datalake
# -------------------------------------------------------------------------
# !env from databricks secrets
CONNECTION_STRING = dbutils.secrets.get(scope='AzureDataLake', key="DATALAKE_CONNECTION_STRING")

# COMMAND ----------

# MAGIC %run /Shared/databricks/au-azure-databricks-cicd/functions/dbrks_helper_functions

# COMMAND ----------

# Load JSON config from Azure datalake
# -------------------------------------------------------------------------
file_path_config = "/config/pipelines/nhsx-au-analytics/"
file_name_config = "config_shared_care_record_dbrks.json"
file_system_config = dbutils.secrets.get(scope='AzureDataLake', key="DATALAKE_CONTAINER_NAME")
config_JSON = datalake_download(CONNECTION_STRING, file_system_config, file_path_config, file_name_config)
config_JSON = json.loads(io.BytesIO(config_JSON).read())

# COMMAND ----------

# Read general parameters from JSON config
# -------------------------------------------------------------------------
file_system = dbutils.secrets.get(scope='AzureDataLake', key="DATALAKE_CONTAINER_NAME")
source_path = config_JSON['pipeline']['raw']['source_path']
sink_path = config_JSON['pipeline']['raw']['sink_path']
historical_source_path = config_JSON['pipeline']['raw']['sink_path']

# COMMAND ----------

# Read org parameters from JSON config, store in dictionaries
# -------------------------------------------------------------------------

table_name_dict = {
'icb': config_JSON['pipeline']['staging'][0]['sink_table'],
'trust': config_JSON['pipeline']['staging'][1]['sink_table'],
'pcn': config_JSON['pipeline']['staging'][2]['sink_table'],
'la': config_JSON['pipeline']['staging'][3]['sink_table'],
'community': config_JSON['pipeline']['staging'][4]['sink_table'],
'other': config_JSON['pipeline']['staging'][5]['sink_table']
}

file_name_dict = {
'icb': config_JSON['pipeline']['proc'][0]['sink_file'],
'trust': config_JSON['pipeline']['proc'][1]['sink_file'],
'pcn': config_JSON['pipeline']['proc'][2]['sink_file'],
'la': config_JSON['pipeline']['proc'][3]['sink_file'],
'community': config_JSON['pipeline']['proc'][4]['sink_file'],
'other': config_JSON['pipeline']['proc'][5]['sink_file'],
}

# COMMAND ----------

# Functions required for data ingestion and processing 
# ----------------------------------------------------

def datalake_listDirectory(CONNECTION_STRING, file_system, source_path):
    try:
        service_client = DataLakeServiceClient.from_connection_string(CONNECTION_STRING)
        file_system_client = service_client.get_file_system_client(file_system=file_system)
        paths = file_system_client.get_paths(path=source_path)
        directory = []
        for path in paths:
            path.name = path.name.replace(source_path, '')
            directory.append(path.name)
    except Exception as e:
        print(e)
    return directory, paths
  

def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames

# COMMAND ----------

# Get LatestFolder and list of all files in latest folder
# --------------------------------------------

latestFolder = datalake_latestFolder(CONNECTION_STRING, file_system, source_path)
directory, paths = datalake_listDirectory(CONNECTION_STRING, file_system, source_path+latestFolder)

# COMMAND ----------

# Ingestion and processing of data from individual excel sheets.
# -------------------------------------------------------------

#initialize org dictionary needed to map variable names to spreadsheet tab name
org_dict = {
'icb': 'ICB',
'trust': 'Trust',
'pcn':'PCN',
'la':'LA',
'community':'Other Community',
'other':'Other partners'
}

#create a dictionary where keys are the  names of the organsiations and the values are blank dataframes that will be popualted later
df_dict = {i:pd.DataFrame() for i in org_dict}

#loop through each submitted file in the landing area. For each file go through the sheets and add append the relevant data to dataframes in df_dict
for filename in directory:
    file = datalake_download(CONNECTION_STRING, file_system, source_path + latestFolder, filename)
    print(filename)

    #Read current file into an iobytes object, read that object and get list of sheet names
    sheets = get_sheetnames_xlsx(io.BytesIO(file))

    ##### ICB CALCULATIONS ICB sheets are different from the other organisations, and so are processed separately ##### 

    #list comprehension to get sheets with ICB in the name from list of all sheets - should only ever be 1 sheet
    sheet_name = [sheet for sheet in sheets if sheet.startswith("ICB")] 
    #Read ICB sheet. The output is a dictionary. Top level is ICB Sheet, next level is each column on the sheet
    xls_file = pd.read_excel(io.BytesIO(file), sheet_name=sheet_name, engine="openpyxl")

    for key in xls_file:
        #drop unnamed columns
        xls_file[key].drop(list(xls_file[key].filter(regex="Unnamed:")), axis=1, inplace=True)
        #remove empty rows if column 1 is empty
        xls_file[key] = xls_file[key].loc[~xls_file[key]["For Month\nsee guidance Ref 2"].isnull()]
        #rename columns based on order they appear in
        xls_file[key].rename(
            columns={
                list(xls_file[key])[0]: "For Month",
                list(xls_file[key])[1]: "ICB ODS code",
                list(xls_file[key])[2]: "ICB Name (if applicable)",
                list(xls_file[key])[3]: "ShCR Programme Name",
                list(xls_file[key])[4]: "Name of ShCR System",
                ###Gap in used columns
                list(xls_file[key])[5]: "Care Providers",
                list(xls_file[key])[6]: "Access to Advanced (EoL) Care Plans",
                #list(xls_file[key])[10]: "Number of users with access to the ShCR",
                list(xls_file[key])[7]: "Number of ShCR views in the past month",
                list(xls_file[key])[8]: "Number of unique user ShCR views in the past month",
                list(xls_file[key])[9]: "Completed by (email)",
                list(xls_file[key])[10]: "Date completed",
            },
            inplace=True,
        )
        
        #users dropped this column from the template. Keeping in to prevent errors but setting to zero
        xls_file[key]["Number of users with access to the ShCR"] = 0

        # get excel file metadata
        ICB_code = xls_file[key]["ICB ODS code"].unique()[0]  # grab ICB code to add in to the other sheets in next section
        ICB_name = xls_file[key]["ICB Name (if applicable)"].unique()[0]  # grab ICB name to add in to the other sheets in next section
          
        #For numeric fields, fill in blanks with zeros. Replace any non numeric entries with zero.
        #xls_file[key]["Number of users with access to the ShCR"] = pd.to_numeric(xls_file[key]["Number of users with access to the ShCR"], errors='coerce').fillna(0).astype(int)
        xls_file[key]["Number of ShCR views in the past month"] = pd.to_numeric(xls_file[key]["Number of ShCR views in the past month"], errors='coerce').fillna(0).astype(int)
        xls_file[key]["Number of unique user ShCR views in the past month"] = pd.to_numeric(xls_file[key]["Number of unique user ShCR views in the past month"], errors='coerce').fillna(0).astype(int)
        xls_file[key]["Care Providers"] = pd.to_numeric(xls_file[key]["Care Providers"], errors='coerce').fillna(0).astype(int)

        # append results to ICB dataframe in df_dict
        df_dict['icb'] = df_dict['icb'].append(xls_file[key], ignore_index=True)


    #### OTHER ORG CALCULATIONS Orgs other than ICB are all the same so can be processed by looping through the dictionary ####
    #get names of other orgs (ignore ICB). Use org dict to map key used here to excel sheet name (e.e. icb to ICB)
    for i in list(df_dict.keys())[1:]:

      sheet_name = [sheet for sheet in sheets if sheet.startswith(org_dict[i])]
      xls_file = pd.read_excel(io.BytesIO(file), sheet_name=sheet_name, engine="openpyxl")
      for key in xls_file:
          
        xls_file[key].drop(list(xls_file[key].filter(regex="Unnamed:")), axis=1, inplace=True)
        xls_file[key] = xls_file[key].loc[~xls_file[key]["For Month"].isnull()]

        xls_file[key].rename(
            columns={
                list(xls_file[key])[0]: "For Month",
                list(xls_file[key])[1]: f"ODS {i} Code",
                list(xls_file[key])[2]: f"{i} Name",
                list(xls_file[key])[3]: "Partner Organisation connected to ShCR?",
                #extra column
                list(xls_file[key])[5]: "Partner Organisation plans to be connected by March 2023?",
            },
            inplace=True,
        )
        #have to use error handling for this one, as column only exists on certain templates
        try: xls_file[key].rename(columns={list(xls_file[key])[6]:'Partner Type'}, inplace= True)
        except: pass

        #add info about the ICB each partner is connected to
        xls_file[key].insert(1, "ICB ODS code", ICB_code, False)
        xls_file[key].insert(3, "ICS Name (if applicable)", ICB_name, False)

        #convert text responses to binary response
        xls_file[key]["Partner Organisation connected to ShCR?"] = xls_file[key]["Partner Organisation connected to ShCR?"].map({"Connected": 1, "Not Connected": 0, "Please select": 0}).fillna(0).astype(int)     
        xls_file[key]["Partner Organisation plans to be connected by March 2023?"] = xls_file[key]["Partner Organisation plans to be connected by March 2023?"].map({"yes": 1, "no": 0, "Yes": 1, "No": 0}).fillna(0).astype(int)
       
          
        # append results to relevant dataframe in df_dict
        df_dict[i] = df_dict[i].append(xls_file[key].iloc[:, 0:9], ignore_index=True)

#Remove any non-required columns from  dataframes
for i in list(df_dict.keys()):
    if i == 'icb': df_dict[i] = df_dict[i][['For Month', 'ICB ODS code', 'ICB Name (if applicable)', 'ShCR Programme Name', 'Name of ShCR System', "Care Providers", 'Access to Advanced (EoL) Care Plans', 'Number of users with access to the ShCR', 'Number of ShCR views in the past month', 'Number of unique user ShCR views in the past month', 'Completed by (email)', 'Date completed']]
    elif i == 'other': df_dict[i] = df_dict[i][['For Month', 'ICB ODS code', 'ICS Name (if applicable)', f'ODS {i} Code', f'{i} Name', 'Partner Organisation connected to ShCR?', 'Partner Organisation plans to be connected by March 2023?', 'Partner Type']]
    else:  df_dict[i] = df_dict[i][['For Month', 'ICB ODS code', 'ICS Name (if applicable)', f'ODS {i} Code', f'{i} Name', 'Partner Organisation connected to ShCR?', 'Partner Organisation plans to be connected by March 2023?']]


# COMMAND ----------

# #cast all date fields to datetime format and set to 1st of the month
#dt =lambda dt: dt.replace(day=1)

#pcn_df['For Month'] = pd.to_datetime(pcn_df['For Month']).apply(dt)
#icb_df['For Month'] = pd.to_datetime(icb_df['For Month']).apply(dt)
#trust_df['For Month'] = pd.to_datetime(trust_df['For Month']).apply(dt)

# COMMAND ----------

#Set all 'For Month' dates to folder date minus 1 month
#--------------------------------------------------
folder_date = pd.to_datetime(latestFolder) - pd.DateOffset(months=1)
df_dict['icb']['For Month'] = folder_date


folder_date = pd.to_datetime(latestFolder)
for i in list(df_dict.keys())[1:]:
  df_dict[i]['For Month'] = folder_date


# COMMAND ----------

#Check for duplicates orgs
#--------------------------------------------------
#create a dictionary of dataframes that contain duplicate entries for each organisation
dupes_dict = {}

for i in df_dict.keys():
  if i == 'icb':
    dupes = [item for item, count in collections.Counter(df_dict[i].iloc[:,1]).items() if count > 1]
    dupes_df = df_dict[i][df_dict[i].iloc[:,1].isin(dupes)]
    dupes_df = dupes_df.iloc[:,[1,2]]
  else:
    dupes = [item for item, count in collections.Counter(df_dict[i].iloc[:,3]).items() if count > 1]
    dupes_df = df_dict[i][df_dict[i].iloc[:,3].isin(dupes)]
    dupes_df = dupes_df.iloc[:,[1,2,3,4]]

  dupes_dict[i] = dupes_df


# COMMAND ----------

#calculate aggregate numbers for all except ICB
#--------------------------------------------------

count_dict = {}

for i in list(df_dict.keys())[1:]:
  count_df = df_dict[i].groupby(df_dict[i].iloc[:,2]).agg(Total = ('Partner Organisation connected to ShCR?', 'size'), Connected = ('Partner Organisation connected to ShCR?', 'sum')).reset_index()
  count_df['Percent'] = count_df['Connected']/count_df['Total']
  count_dict[i] = count_df


# COMMAND ----------

#SNAPSHOT SUMMARY
#Write pages to Excel file in iobytes
#--------------------------------------------------

files = []
sheets = []

#loop though main, counts and dupes dataframes and add to list
for i in df_dict.keys():
  files.append(df_dict[i])
  if i !='icb':
    files.append(count_dict[i])
  files.append(dupes_dict[i])

#loop though main, counts and dupes dataframe keys and add to list
for i in df_dict.keys():
  sheets.append(i)
  if i !='icb':
    sheets.append(f'{i} Count')
  sheets.append(f'{i} Dupes')

excel_sheet = io.BytesIO()

#write files to excel file, using sheets to name excel sheets
writer = pd.ExcelWriter(excel_sheet, engine='openpyxl')
for count, file in enumerate(files):
   file.to_excel(writer, sheet_name=sheets[count], index=False)
writer.save()

# COMMAND ----------

#SNAPSHOT SUMMARY
#Send Excel snapshot File to test Output in datalake
#--------------------------------------------------

current_date_path = datetime.now().strftime('%Y-%m-%d') + '/'
file_contents = excel_sheet
datalake_upload(file_contents, CONNECTION_STRING, file_system, "proc/projects/nhsx_slt_analytics/shcr/excel_summary/"+current_date_path, "shared_care_summary_output.xlsx")

# COMMAND ----------

#drop partner type from other dataframe once it has been included in output summary
df_dict['other'].drop(columns=('Partner Type'), inplace = True)

# COMMAND ----------

#Pull historical files
#------------------------------------

latestFolder = datalake_latestFolder(CONNECTION_STRING, file_system, historical_source_path)
historic_file_list = datalake_listContents(CONNECTION_STRING, file_system, historical_source_path+latestFolder)

historic_df_dict = {}

#read files in latest folder, loop through possible organisation names to determine what to key value to assign to dataframe
for file in historic_file_list:
  for i in df_dict.keys():
    if i in file:
      historic_parquet = datalake_download(CONNECTION_STRING, file_system, historical_source_path+latestFolder, file)
      print(file)
      historic_df = pd.read_parquet(io.BytesIO(historic_parquet), engine="pyarrow")
      historic_df['For Month'] = pd.to_datetime(historic_df['For Month'])
      if i =='icb':
        pass
        historic_df['Date completed'] = pd.to_datetime(historic_df['Date completed'],errors='coerce')
      historic_df_dict[i] = historic_df

# COMMAND ----------

# #CODE TO CREATE INITIAL HISTORICAL FILES (DONT USE UNLESS STARTING DATA)

# current_date_path = datetime.now().strftime('%Y-%m-%d') + '/'
# # # PCN
# # # #-----------------
# file_contents = io.BytesIO()
# df_dict['pcn'] = df_dict['pcn'].astype(str)
# df_dict['pcn'].to_parquet(file_contents, engine="pyarrow")
# datalake_upload(file_contents, CONNECTION_STRING, file_system, sink_path+current_date_path, "shcr_partners_pcn_data_month_count.parquet")

# # ICB
# #-----------------
# file_contents = io.BytesIO()
# df_dict['icb'] = df_dict['icb'].astype(str)
# df_dict['icb'] = df_dict['icb'].astype(str)
# df_dict['icb'].to_parquet(file_contents, engine="pyarrow")
# datalake_upload(file_contents, CONNECTION_STRING, file_system, sink_path+current_date_path, "shcr_partners_icb_data_month_count.parquet")

# # NHS Trust
# #-----------------
# file_contents = io.BytesIO()
# df_dict['trust'] = df_dict['trust'].astype(str)
# df_dict['trust'].to_parquet(file_contents, engine="pyarrow")
# datalake_upload(file_contents, CONNECTION_STRING, file_system, sink_path+current_date_path, "shcr_partners_trust_data_month_count.parquet")

# # NHS LA
# #-----------------
# file_contents = io.BytesIO()
# df_dict['la'] = df_dict['la'].astype(str)
# df_dict['la'].to_parquet(file_contents, engine="pyarrow")
# datalake_upload(file_contents, CONNECTION_STRING, file_system, sink_path+current_date_path, "shcr_partners_la_data_month_count.parquet")

# # # NHS Community
# # #-----------------
# file_contents = io.BytesIO()
# df_dict['community'] = df_dict['community'].astype(str)
# df_dict['community'].to_parquet(file_contents, engine="pyarrow")
# datalake_upload(file_contents, CONNECTION_STRING, file_system, sink_path+current_date_path, "shcr_partners_community_data_month_count.parquet")

# # # # NHS Other
# # # #-----------------
# file_contents = io.BytesIO()
# df_dict['other'] = df_dict['other'].astype(str)
# df_dict['other'].to_parquet(file_contents, engine="pyarrow")
# datalake_upload(file_contents, CONNECTION_STRING, file_system, sink_path+current_date_path, "shcr_partners_other_data_month_count.parquet")

# COMMAND ----------

# Append new data to historical data
#-----------------------------------------------------------------------
for i in df_dict.keys():
  print(i)
  dates_in_historic = historic_df_dict[i]["For Month"].unique().tolist()
  #takes the first date in the new data (will only be one) and checks if it exists in historic data
  dates_in_new = df_dict[i]["For Month"].unique().tolist()[0]
  if dates_in_new in dates_in_historic:
    print(f'{i} Data already exists in historical data')
  else:
    historic_df_dict[i] = historic_df_dict[i].append(df_dict[i])
    #historic_df_dict[i] = historic_df_dict[i].sort_values(by=['For Month'])
    historic_df_dict[i] = historic_df_dict[i].reset_index(drop=True)
    historic_df_dict[i] = historic_df_dict[i].astype(str)


for i in df_dict.keys():
  historic_df_dict[i].reset_index(drop = True)
  historic_df_dict[i].index.name = "Unique ID"

# COMMAND ----------

#drop rows where org name is nan (this should catch columns where the ICB hasn't entered anything)
for i in list(df_dict.keys())[1:]:
  column = f'{i} Name'
  historic_df_dict[i] = historic_df_dict[i].drop(historic_df_dict[i][historic_df_dict[i][column]=='nan'].index)

# COMMAND ----------

# Upload processed data to datalake
#-----------------------------------------------------------------------

current_date_path = datetime.now().strftime('%Y-%m-%d') + '/'

for i in historic_df_dict.keys():
  print(i)
  file_contents = io.BytesIO()
  historic_df_dict[i] = historic_df_dict[i].astype(str)
  historic_df_dict[i].to_parquet(file_contents, engine="pyarrow")
  filename = file_name_dict[i]
  datalake_upload(file_contents, CONNECTION_STRING, file_system, sink_path+current_date_path, filename)

# COMMAND ----------

#convert ICB columns to numeric
icb_numeric_columns = ['Care Providers', 'Number of users with access to the ShCR', 'Number of ShCR views in the past month', 'Number of unique user ShCR views in the past month']
historic_df_dict['icb'][icb_numeric_columns] = historic_df_dict['icb'][icb_numeric_columns].apply(pd.to_numeric)

#convert other org columns to numeric
part_numeric_columns = ['Partner Organisation connected to ShCR?', 'Partner Organisation plans to be connected by March 2023?']
for i in list(df_dict.keys())[1:]:
  historic_df_dict[i][part_numeric_columns] = historic_df_dict[i][part_numeric_columns].apply(pd.to_numeric)

# COMMAND ----------

for i in historic_df_dict.keys():
  write_to_sql(historic_df_dict[i], table_name_dict[i], "overwrite")
  print(table_name_dict[i])
